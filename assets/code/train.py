import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader
from dataset import EmotionDataset
from model import EmotionCNN
import numpy as np
from sklearn.metrics import f1_score, precision_score, confusion_matrix
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 参数设置
BATCH_SIZE = 32
EPOCHS = 60
LEARNING_RATE = 0.001
EMOTION_NAMES = ["angry", "disgust", "fear", "happy", "sad", "surprise", "neutral"]


# 训练一个epoch
def train_epoch(model, loader, criterion, optimizer, device):
    model.train()
    total_loss = 0
    all_preds = []
    all_labels = []

    for images, labels in loader:
        images = images.to(device)
        labels = labels.to(device)

        optimizer.zero_grad()
        outputs = model(images)
        loss = criterion(outputs, labels)
        loss.backward()
        optimizer.step()

        total_loss += loss.item()
        _, predicted = outputs.max(1)
        all_preds.extend(predicted.cpu().numpy())
        all_labels.extend(labels.cpu().numpy())

    avg_loss = total_loss / len(loader)
    acc = 100.0 * np.mean(np.array(all_preds) == np.array(all_labels))
    f1 = f1_score(all_labels, all_preds, average="weighted", zero_division=0)
    precision = precision_score(all_labels, all_preds, average="weighted", zero_division=0)
    rmse = compute_rmse(all_labels, all_preds)

    return avg_loss, acc, f1, precision, rmse

# 计算RMSE
def compute_rmse(labels, preds):
    labels = np.array(labels)
    preds = np.array(preds)
    rmse = np.sqrt(np.mean((preds - labels) ** 2))
    return rmse
# 验证
def validate(model, loader, criterion, device):
    model.eval()
    total_loss = 0
    all_preds = []
    all_labels = []

    with torch.no_grad():
        for images, labels in loader:
            images = images.to(device)
            labels = labels.to(device)

            outputs = model(images)
            loss = criterion(outputs, labels)
            total_loss += loss.item()

            _, predicted = outputs.max(1)
            all_preds.extend(predicted.cpu().numpy())
            all_labels.extend(labels.cpu().numpy())

    avg_loss = total_loss / len(loader)
    acc = 100.0 * np.mean(np.array(all_preds) == np.array(all_labels))
    f1 = f1_score(all_labels, all_preds, average="weighted", zero_division=0)
    precision = precision_score(all_labels, all_preds, average="weighted", zero_division=0)
    rmse = compute_rmse(all_labels, all_preds)

    # 每个类别单独的f1和precision
    f1_each = f1_score(all_labels, all_preds, average=None, zero_division=0)
    precision_each = precision_score(all_labels, all_preds, average=None, zero_division=0)
    cm = confusion_matrix(all_labels, all_preds)

    return avg_loss, acc, f1, precision, rmse, f1_each, precision_each, cm


# 保存结果到excel
def save_results_to_excel(history, best_f1_each, best_precision_each, best_cm, best_acc):
    os.makedirs("models", exist_ok=True)
    save_path = "models/training_results.xlsx"

    wb = openpyxl.Workbook()

    # 样式
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    def make_border():
        thin = Side(style="thin")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    # 颜色
    COLOR_BLUE = "4472C4"
    COLOR_GRAY = "F2F2F2"
    COLOR_GREEN = "E2EFDA"
    COLOR_YELLOW = "FFF2CC"

    # 表头样式
    def set_header(ws, row, col, text, bg=None):
        if bg is None:
            bg = COLOR_BLUE
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = center
        cell.border = make_border()
        return cell

    # 数据单元格样式
    def set_cell(ws, row, col, value, align=None, bold=False, bg=None, num_format=None):
        if align is None:
            align = center
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=bold, name="Calibri", size=10)
        cell.alignment = align
        cell.border = make_border()
        if bg is not None:
            cell.fill = PatternFill("solid", fgColor=bg)
        if num_format is not None:
            cell.number_format = num_format
        return cell

    # 设置列宽和冻结窗格
    def set_layout(ws, freeze, col_widths):
        ws.freeze_panes = freeze
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
        ws.row_dimensions[1].height = 22

    # Sheet1: 训练历史
    ws1 = wb.active
    ws1.title = "训练历史"

    headers = [
        "Epoch",
        "Train Loss", "Train ACC (%)", "Train F1", "Train Precision", "Train RMSE",
        "Val Loss", "Val ACC (%)", "Val F1", "Val Precision", "Val RMSE",
        "最佳模型"
    ]
    for col, h in enumerate(headers, 1):
        set_header(ws1, 1, col, h)

    for row_idx, record in enumerate(history, 2):
        is_best = record["is_best"]

        # 最佳模型行用绿色，其余行间隔灰色
        if is_best:
            bg = COLOR_GREEN
        elif row_idx % 2 == 0:
            bg = COLOR_GRAY
        else:
            bg = None

        values = [
            record["epoch"],
            record["train_loss"], record["train_acc"],
            record["train_f1"], record["train_prec"], record["train_rmse"],
            record["val_loss"], record["val_acc"],
            record["val_f1"], record["val_prec"], record["val_rmse"],
            "✓" if is_best else ""
        ]
        formats = [
            None,
            "0.0000", "0.00", "0.0000", "0.0000", "0.0000",
            "0.0000", "0.00", "0.0000", "0.0000", "0.0000",
            None
        ]
        for col, (val, fmt) in enumerate(zip(values, formats), 1):
            set_cell(ws1, row_idx, col, val, bg=bg, num_format=fmt, bold=is_best)

    set_layout(ws1, "B2", {
        "A": 8, "B": 12, "C": 14, "D": 12, "E": 16, "F": 12,
        "G": 12, "H": 12, "I": 10, "J": 16, "K": 12, "L": 10
    })

    # Sheet2: 各类别指标
    ws2 = wb.create_sheet("各类别指标（最佳模型）")

    for col, h in enumerate(["情绪类别", "F1 Score", "Precision"], 1):
        set_header(ws2, 1, col, h)

    for i, name in enumerate(EMOTION_NAMES):
        row_idx = i + 2
        bg = COLOR_GRAY if row_idx % 2 == 0 else None
        set_cell(ws2, row_idx, 1, name, align=left, bg=bg)
        set_cell(ws2, row_idx, 2, round(float(best_f1_each[i]), 4), bg=bg, num_format="0.0000")
        set_cell(ws2, row_idx, 3, round(float(best_precision_each[i]), 4), bg=bg, num_format="0.0000")

    # 加一行整体平均
    total_row = len(EMOTION_NAMES) + 3
    set_cell(ws2, total_row, 1, "整体 (weighted)", align=left, bold=True, bg=COLOR_YELLOW)
    set_cell(ws2, total_row, 2, round(float(best_f1_each.mean()), 4), bold=True, bg=COLOR_YELLOW, num_format="0.0000")
    set_cell(ws2, total_row, 3, round(float(best_precision_each.mean()), 4), bold=True, bg=COLOR_YELLOW, num_format="0.0000")

    set_layout(ws2, "B2", {"A": 20, "B": 12, "C": 12})

    # Sheet3: 混淆矩阵
    ws3 = wb.create_sheet("混淆矩阵（最佳模型）")

    # 左上角
    corner = ws3.cell(row=1, column=1, value="真实 \\ 预测")
    corner.font = Font(bold=True, name="Calibri", size=10, italic=True)
    corner.alignment = center
    corner.border = make_border()
    corner.fill = PatternFill("solid", fgColor="D9D9D9")

    # 列标题
    for col, name in enumerate(EMOTION_NAMES, 2):
        set_header(ws3, 1, col, name)

    # 数据
    row_totals = best_cm.sum(axis=1)
    for row_idx, name in enumerate(EMOTION_NAMES, 2):
        set_header(ws3, row_idx, 1, name)
        for col, val in enumerate(best_cm[row_idx - 2], 2):
            is_diagonal = (col - 2 == row_idx - 2)
            percent = val / row_totals[row_idx - 2] * 100 if row_totals[row_idx - 2] > 0 else 0

            # 对角线绿色（预测正确），高错误率红色，中等错误率黄色
            if is_diagonal:
                bg = "C6EFCE"
            elif percent > 20:
                bg = "FFC7CE"
            elif percent > 10:
                bg = "FFEB9C"
            else:
                bg = None

            cell = ws3.cell(row=row_idx, column=col, value=int(val))
            cell.font = Font(
                bold=is_diagonal,
                name="Calibri",
                size=10,
                color="375623" if is_diagonal else "000000"
            )
            cell.alignment = center
            cell.border = make_border()
            if bg is not None:
                cell.fill = PatternFill("solid", fgColor=bg)

    set_layout(ws3, "B2", {
        "A": 14, "B": 10, "C": 10, "D": 10,
        "E": 10, "F": 10, "G": 10, "H": 10
    })
    for i in range(1, len(EMOTION_NAMES) + 2):
        ws3.row_dimensions[i].height = 20

    # Sheet4: 汇总信息 
    ws4 = wb.create_sheet("汇总")
    set_header(ws4, 1, 1, "项目")
    set_header(ws4, 1, 2, "值")

    summary_data = [
        ("最佳验证准确率", f"{best_acc:.2f}%"),
        ("总训练轮数", EPOCHS),
        ("Batch Size", BATCH_SIZE),
        ("Learning Rate", LEARNING_RATE),
        ("模型保存路径", "models/best_model.pth"),
        ("结果保存路径", "models/training_results.xlsx"),
    ]
    for row_idx, (key, value) in enumerate(summary_data, 2):
        bg = COLOR_GRAY if row_idx % 2 == 0 else None
        set_cell(ws4, row_idx, 1, key, align=left, bg=bg)
        set_cell(ws4, row_idx, 2, value, align=left, bg=bg)

    set_layout(ws4, "A2", {"A": 22, "B": 30})

    wb.save(save_path)
    print(f"\n结果已保存到: {save_path}")


# 主函数
def main():
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"使用设备: {device}")

    # 加载数据集
    train_dataset = EmotionDataset("data/processed/train", is_train=True)
    val_dataset = EmotionDataset("data/processed/val", is_train=False)
    train_loader = DataLoader(train_dataset, batch_size=BATCH_SIZE, shuffle=True)
    val_loader = DataLoader(val_dataset, batch_size=BATCH_SIZE, shuffle=False)

    # 创建模型
    model = EmotionCNN().to(device)

    # 自动统计训练集中各个类别的样本数量，计算真实客观的权重
    labels = [sample[1] for sample in train_dataset.samples] # 获取所有训练标签
    class_counts = np.bincount(labels) # 统计每个标签出现的次数
    total_samples = len(labels)
    
    # 权重计算：总样本数 / (类别数 * 该类样本数)，样本越少则权重越高
    weights = total_samples / (len(EMOTION_NAMES) * class_counts)
    class_weights = torch.tensor(weights, dtype=torch.float).to(device)
    
    print(f"自动计算的类别权重为: {class_weights}")
    criterion = nn.CrossEntropyLoss(weight=class_weights)

    optimizer = optim.Adam(model.parameters(), lr=LEARNING_RATE)

    # 将调度器换为 ReduceLROnPlateau
    scheduler = optim.lr_scheduler.ReduceLROnPlateau(optimizer, mode='max', factor=0.5, patience=5)

    # 训练循环内部将 step() 放在 validate 之后：
    val_loss, val_acc, val_f1, val_prec, val_rmse, f1_each, precision_each, cm = validate(model, val_loader, criterion, device)
    scheduler.step(val_acc) # 依据验证集准确率决定是否降低学习率

    # 记录训练过程
    best_acc = 0
    history = []
    best_f1_each = None
    best_precision_each = None
    best_cm = None

    for epoch in range(EPOCHS):
        # 训练
        train_loss, train_acc, train_f1, train_prec, train_rmse = train_epoch(
            model, train_loader, criterion, optimizer, device
        )

        # 验证
        val_loss, val_acc, val_f1, val_prec, val_rmse, f1_each, precision_each, cm = validate(
            model, val_loader, criterion, device
        )

        # 更新学习率 (使用 ReduceLROnPlateau 的新方法)
        scheduler.step(val_acc)  # 把验证集准确率传进去，它才知道什么时候该降学习率
        current_lr = optimizer.param_groups[0]['lr']  # ReduceLROnPlateau 没有 get_last_lr() 方法，必须从优化器里直接读

        # 如果这次验证准确率更高，就保存模型
        is_best = val_acc > best_acc
        if is_best:
            best_acc = val_acc
            best_f1_each = f1_each.copy()
            best_precision_each = precision_each.copy()
            best_cm = cm.copy()
            os.makedirs("models", exist_ok=True)
            torch.save(model.state_dict(), "models/best_model.pth")

        # 记录这个epoch的结果
        history.append({
            "epoch": epoch + 1,
            "train_loss": round(train_loss, 4),
            "train_acc": round(train_acc, 2),
            "train_f1": round(train_f1, 4),
            "train_prec": round(train_prec, 4),
            "train_rmse": round(train_rmse, 4),
            "val_loss": round(val_loss, 4),
            "val_acc": round(val_acc, 2),
            "val_f1": round(val_f1, 4),
            "val_prec": round(val_prec, 4),
            "val_rmse": round(val_rmse, 4),
            "is_best": is_best,
        })

        # 打印结果
        print(f"\nEpoch {epoch + 1}/{EPOCHS}  (当前学习率: {current_lr:.6f})")
        print(f"  训练集 -> Loss: {train_loss:.4f} | ACC: {train_acc:.2f}% | F1: {train_f1:.4f} | Precision: {train_prec:.4f}")
        print(f"  验证集 -> Loss: {val_loss:.4f} | ACC: {val_acc:.2f}% | F1: {val_f1:.4f} | Precision: {val_prec:.4f}")
        print("  各类别验证结果:")
        for i, name in enumerate(EMOTION_NAMES):
            print(f"    {name:10s} F1: {f1_each[i]:.4f}  Precision: {precision_each[i]:.4f}")
        if is_best:
            print(f"  *** 保存最佳模型，当前最高准确率: {val_acc:.2f}% ***")

    # 训练结束，打印最终结果
    print(f"\n训练完成！最佳验证准确率: {best_acc:.2f}%")
    print("\n混淆矩阵（最佳模型）:")
    header = "真实\\预测  " + "  ".join(f"{n:>8}" for n in EMOTION_NAMES)
    print(header)
    print("-" * len(header))
    for i, row in enumerate(best_cm):
        print(f"{EMOTION_NAMES[i]:10s}" + "  ".join(f"{v:8d}" for v in row))

    # 保存到excel
    save_results_to_excel(history, best_f1_each, best_precision_each, best_cm, best_acc)


if __name__ == "__main__":
    main()